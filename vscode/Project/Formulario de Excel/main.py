import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import re
import os

nombre_archivo = "datos.xlsx"

# Verificar si el archivo ya existe y abrirlo o crear uno nuevo
if os.path.exists(nombre_archivo):
    wb = load_workbook(nombre_archivo)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Nombre", "Edad", "Email", "Teléfono", "Dirección"])
    wb.save(nombre_archivo)  # Guardar el archivo nuevo inmediatamente


def guardar_datos():
    nombre = entry_nombre.get().strip()
    edad = entry_edad.get().strip()
    email = entry_email.get().strip()
    telefono = entry_telefono.get().strip()
    direccion = entry_direccion.get().strip()

    # Validación: No permitir campos vacíos
    if not nombre or not edad or not email or not telefono or not direccion:
        messagebox.showwarning(
            "Advertencia", "Todos los campos son obligatorios")
        return

    try:
        edad = int(edad)
        telefono = int(telefono)
    except ValueError:
        messagebox.showerror("Error", "Edad y Teléfono deben ser números")
        return

    # Validación mejorada para el correo electrónico
    if not re.match(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$", email):
        messagebox.showwarning(
            "Advertencia", "El correo electrónico no es válido")
        return

    # Guardar los datos en Excel
    ws.append([nombre, edad, email, telefono, direccion])
    wb.save(nombre_archivo)  # Asegurar que se guarden los cambios

    # Mostrar mensaje de éxito
    messagebox.showinfo("Información", "Datos guardados con éxito")

    # Borrar los campos de entrada correctamente
    entry_nombre.delete(0, tk.END)
    entry_edad.delete(0, tk.END)
    entry_email.delete(0, tk.END)
    entry_telefono.delete(0, tk.END)
    entry_direccion.delete(0, tk.END)


# Configuración de la ventana
root = tk.Tk()
root.title("Formulario de Entrada de Datos")
root.configure(bg="#2e4053")

label_style = {"bg": "#2e4053", "fg": "white"}
entry_style = {"bg": "#eaecee", "fg": "black"}

# Etiquetas y Entradas
label_nombre = tk.Label(root, text="Nombre", **label_style)
label_nombre.grid(row=0, column=0, padx=10, pady=5)
entry_nombre = tk.Entry(root, **entry_style)
entry_nombre.grid(row=0, column=1, padx=10, pady=5)

label_edad = tk.Label(root, text="Edad", **label_style)
label_edad.grid(row=1, column=0, padx=10, pady=5)
entry_edad = tk.Entry(root, **entry_style)
entry_edad.grid(row=1, column=1, padx=10, pady=5)

label_email = tk.Label(root, text="Email", **label_style)
label_email.grid(row=2, column=0, padx=10, pady=5)
entry_email = tk.Entry(root, **entry_style)
entry_email.grid(row=2, column=1, padx=10, pady=5)

label_telefono = tk.Label(root, text="Teléfono", **label_style)
label_telefono.grid(row=3, column=0, padx=10, pady=5)
entry_telefono = tk.Entry(root, **entry_style)
entry_telefono.grid(row=3, column=1, padx=10, pady=5)

label_direccion = tk.Label(root, text="Dirección", **label_style)
label_direccion.grid(row=4, column=0, padx=10, pady=5)
entry_direccion = tk.Entry(root, **entry_style)
entry_direccion.grid(row=4, column=1, padx=10, pady=5)

# Botón de Guardar
boton_guardar = tk.Button(
    root, text="Guardar", command=guardar_datos, bg="#85929e", fg="white", width=30)
boton_guardar.grid(row=5, column=0, columnspan=2, padx=10, pady=10)

root.mainloop()
